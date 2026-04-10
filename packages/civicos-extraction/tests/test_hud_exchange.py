"""Tests for HUD Exchange client — allocation parsing, filtering, and storage mapping."""

import time
from unittest.mock import MagicMock, patch, PropertyMock

import pytest
import requests

from civicos_extraction.clients.hud_exchange import (
    COLUMN_MAPPINGS,
    HUD_ALLOCATION_URLS,
    HUD_PROGRAMS,
    HUDAllocation,
    HUDExchangeClient,
    create_hud_exchange_client,
    extract_allocations_to_storage,
    hud_allocation_to_storage,
)


# ==================== Fixtures ====================


@pytest.fixture
def client(tmp_path):
    """HUDExchangeClient with zero delay and temp cache dir."""
    return HUDExchangeClient(
        jurisdiction_id="federal-US",
        cache_dir=str(tmp_path),
        request_delay=0.0,
    )


@pytest.fixture
def sample_allocation():
    """A realistic HUD allocation record."""
    return HUDAllocation(
        grantee_name="City of San Rafael",
        state="CA",
        program="CDBG",
        fiscal_year=2025,
        amount_cents=52300000,
        grantee_type="entitlement",
        source_url="https://www.hud.gov/sites/dfiles/CPD/documents/FY2025-Formula-Allocations-All-Grantees.xlsx",
    )


@pytest.fixture
def sample_allocations_data():
    """Raw allocation dicts as returned by _parse_excel."""
    return [
        {
            "grantee_name": "City of San Rafael",
            "state": "CA",
            "program": "CDBG",
            "fiscal_year": 2025,
            "amount_cents": 52300000,
            "grantee_type": "entitlement",
            "source_url": "https://example.com/fy2025.xlsx",
        },
        {
            "grantee_name": "City of San Rafael",
            "state": "CA",
            "program": "HOME",
            "fiscal_year": 2025,
            "amount_cents": 18700000,
            "grantee_type": "entitlement",
            "source_url": "https://example.com/fy2025.xlsx",
        },
        {
            "grantee_name": "City of Berkeley",
            "state": "CA",
            "program": "CDBG",
            "fiscal_year": 2025,
            "amount_cents": 320000000,
            "grantee_type": "entitlement",
            "source_url": "https://example.com/fy2025.xlsx",
        },
        {
            "grantee_name": "Portland",
            "state": "OR",
            "program": "CDBG",
            "fiscal_year": 2025,
            "amount_cents": 800000000,
            "grantee_type": "entitlement",
            "source_url": "https://example.com/fy2025.xlsx",
        },
    ]


# ==================== HUDAllocation dataclass ====================


class TestHUDAllocation:
    """Tests for the HUDAllocation dataclass."""

    def test_to_dict_returns_all_fields(self, sample_allocation):
        d = sample_allocation.to_dict()
        assert d["grantee_name"] == "City of San Rafael"
        assert d["state"] == "CA"
        assert d["program"] == "CDBG"
        assert d["fiscal_year"] == 2025
        assert d["amount_cents"] == 52300000
        assert d["grantee_type"] == "entitlement"
        assert "FY2025" in d["source_url"]

    def test_to_dict_round_trips_all_keys(self, sample_allocation):
        d = sample_allocation.to_dict()
        expected_keys = {"grantee_name", "state", "program", "fiscal_year",
                         "amount_cents", "grantee_type", "source_url"}
        assert set(d.keys()) == expected_keys


# ==================== Client properties ====================


class TestClientProperties:
    """Tests for HUDExchangeClient identity properties."""

    def test_platform_name(self, client):
        assert client.platform_name == "hud_exchange"

    def test_source_id_includes_jurisdiction(self, client):
        assert client.source_id == "hud_exchange-federal-US"

    def test_source_id_custom_jurisdiction(self, tmp_path):
        c = HUDExchangeClient(jurisdiction_id="city-san-rafael", cache_dir=str(tmp_path))
        assert c.source_id == "hud_exchange-city-san-rafael"

    def test_source_type(self, client):
        assert client.source_type == "hud_exchange"

    def test_default_jurisdiction_id(self, client):
        assert client.jurisdiction_id == "federal-US"


# ==================== _normalize_grantee_name ====================


class TestNormalizeGranteeName:
    """Tests for grantee name normalization logic."""

    def test_strips_city_of_prefix(self, client):
        assert client._normalize_grantee_name("City of San Rafael") == "san rafael"

    def test_strips_town_of_prefix(self, client):
        assert client._normalize_grantee_name("Town of Fairfax") == "fairfax"

    def test_strips_county_of_prefix(self, client):
        assert client._normalize_grantee_name("County of Marin") == "marin"

    def test_strips_village_of_prefix(self, client):
        assert client._normalize_grantee_name("Village of Scarsdale") == "scarsdale"

    def test_lowercases(self, client):
        assert client._normalize_grantee_name("SAN RAFAEL") == "san rafael"

    def test_strips_whitespace(self, client):
        assert client._normalize_grantee_name("  San Rafael  ") == "san rafael"

    def test_no_prefix_passthrough(self, client):
        assert client._normalize_grantee_name("Portland") == "portland"

    def test_empty_string(self, client):
        assert client._normalize_grantee_name("") == ""


# ==================== _infer_grantee_type ====================


class TestInferGranteeType:
    """Tests for sheet-name-based grantee type inference."""

    def test_entitlement_sheet(self, client):
        assert client._infer_grantee_type("Entitlement Grantees") == "entitlement"

    def test_state_sheet(self, client):
        assert client._infer_grantee_type("State Grantees") == "state"

    def test_insular_sheet(self, client):
        assert client._infer_grantee_type("Insular Areas") == "insular_area"

    def test_urban_county_sheet(self, client):
        assert client._infer_grantee_type("Urban County Grantees") == "entitlement"

    def test_consortium_sheet(self, client):
        assert client._infer_grantee_type("Consortium Grantees") == "entitlement"

    def test_unknown_defaults_entitlement(self, client):
        assert client._infer_grantee_type("Sheet1") == "entitlement"

    def test_case_insensitive(self, client):
        assert client._infer_grantee_type("STATE PROGRAMS") == "state"


# ==================== _normalize_grantee_type ====================


class TestNormalizeGranteeType:
    """Tests for row-level grantee type normalization."""

    def test_non_entitlement(self, client):
        assert client._normalize_grantee_type("Non-Entitlement") == "state"

    def test_state_type(self, client):
        assert client._normalize_grantee_type("State Program") == "state"

    def test_principal_city(self, client):
        assert client._normalize_grantee_type("Principal City") == "entitlement"

    def test_metro_city(self, client):
        assert client._normalize_grantee_type("Metro City") == "entitlement"

    def test_urban_county(self, client):
        assert client._normalize_grantee_type("Urban County") == "entitlement"

    def test_consortium(self, client):
        assert client._normalize_grantee_type("Consortium") == "entitlement"

    def test_insular(self, client):
        assert client._normalize_grantee_type("Insular Area") == "insular_area"

    def test_unknown_defaults_entitlement(self, client):
        assert client._normalize_grantee_type("Something Else") == "entitlement"


# ==================== _map_columns ====================


class TestMapColumns:
    """Tests for Excel column header mapping."""

    def test_maps_standard_fy2025_headers(self, client):
        headers = {"NAME": 0, "STA": 1, "Type": 2, "CDBG": 3, "HOME": 4, "ESG": 5}
        col_map = client._map_columns(headers)
        assert col_map["grantee"] == 0
        assert col_map["state"] == 1
        assert col_map["type"] == 2
        assert col_map["cdbg"] == 3
        assert col_map["home"] == 4
        assert col_map["esg"] == 5

    def test_maps_alternate_header_names(self, client):
        headers = {"Grantee Name": 0, "STATE": 1, "CDBG Allocation": 2}
        col_map = client._map_columns(headers)
        assert col_map["grantee"] == 0
        assert col_map["state"] == 1
        assert col_map["cdbg"] == 2

    def test_missing_columns_not_in_map(self, client):
        headers = {"NAME": 0}
        col_map = client._map_columns(headers)
        assert "grantee" in col_map
        assert "cdbg" not in col_map
        assert "home" not in col_map

    def test_empty_headers(self, client):
        col_map = client._map_columns({})
        assert col_map == {}

    def test_case_insensitive_matching(self, client):
        headers = {"name": 0, "cdbg": 1}
        col_map = client._map_columns(headers)
        assert col_map["grantee"] == 0
        assert col_map["cdbg"] == 1

    def test_all_program_columns(self, client):
        headers = {"NAME": 0, "CDBG": 1, "HOME": 2, "ESG": 3, "HOPWA": 4, "HTF": 5, "RHP": 6}
        col_map = client._map_columns(headers)
        for program in ["cdbg", "home", "esg", "hopwa", "htf", "rhp"]:
            assert program in col_map


# ==================== _throttle_request ====================


class TestThrottleRequest:
    """Tests for request throttling."""

    def test_throttle_waits_if_too_fast(self, tmp_path):
        c = HUDExchangeClient(cache_dir=str(tmp_path), request_delay=0.1)
        c.last_request_time = time.time()  # Just requested

        start = time.time()
        c._throttle_request()
        elapsed = time.time() - start

        # Should have waited approximately request_delay seconds
        assert elapsed >= 0.05  # Allow some tolerance

    def test_throttle_skips_if_enough_time_passed(self, tmp_path):
        c = HUDExchangeClient(cache_dir=str(tmp_path), request_delay=0.1)
        c.last_request_time = time.time() - 10.0  # Long ago

        start = time.time()
        c._throttle_request()
        elapsed = time.time() - start

        # Should not wait
        assert elapsed < 0.05

    def test_throttle_updates_last_request_time(self, client):
        old_time = client.last_request_time
        client._throttle_request()
        assert client.last_request_time > old_time


# ==================== _download_excel ====================


class TestDownloadExcel:
    """Tests for Excel file downloading with caching."""

    def test_download_success_saves_file(self, client, tmp_path):
        mock_response = MagicMock()
        mock_response.content = b"fake-xlsx-content"
        mock_response.raise_for_status = MagicMock()

        with patch.object(client.session, "get", return_value=mock_response):
            result = client._download_excel("https://example.com/test.xlsx", 2025)

        assert result is not None
        assert result.endswith("hud_cpd_fy2025.xlsx")
        # Verify actual file was written
        with open(result, "rb") as f:
            assert f.read() == b"fake-xlsx-content"

    def test_download_uses_cache_for_recent_file(self, client, tmp_path):
        # Create a "recent" cached file
        cache_file = tmp_path / "hud_cpd_fy2025.xlsx"
        cache_file.write_bytes(b"cached-content")

        result = client._download_excel("https://example.com/test.xlsx", 2025)

        assert result == str(cache_file)

    def test_download_redownloads_old_cached_file(self, client, tmp_path):
        import os

        # Create an "old" cached file (> 7 days)
        cache_file = tmp_path / "hud_cpd_fy2025.xlsx"
        cache_file.write_bytes(b"old-content")
        old_time = time.time() - 8 * 24 * 60 * 60  # 8 days ago
        os.utime(str(cache_file), (old_time, old_time))

        mock_response = MagicMock()
        mock_response.content = b"fresh-content"
        mock_response.raise_for_status = MagicMock()

        with patch.object(client.session, "get", return_value=mock_response):
            result = client._download_excel("https://example.com/test.xlsx", 2025)

        with open(result, "rb") as f:
            assert f.read() == b"fresh-content"

    def test_download_failure_returns_none(self, client):
        with patch.object(
            client.session, "get",
            side_effect=requests.RequestException("Network error"),
        ):
            result = client._download_excel("https://example.com/test.xlsx", 2025)

        assert result is None


# ==================== _parse_excel ====================


class TestParseExcel:
    """Tests for Excel file parsing using openpyxl."""

    @pytest.fixture
    def simple_xlsx(self, tmp_path):
        """Create a minimal Excel file with allocation data."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Entitlement Grantees"

        # Header row
        ws.append(["NAME", "STA", "Type", "CDBG", "HOME", "ESG"])
        # Data rows
        ws.append(["City of San Rafael", "CA", "Metro City", 523000, 187000, 45000])
        ws.append(["City of Berkeley", "CA", "Principal City", 3200000, 1100000, 250000])
        ws.append(["Portland", "OR", "Metro City", 8000000, None, 600000])

        path = tmp_path / "test_fy2025.xlsx"
        wb.save(str(path))
        return str(path)

    @pytest.fixture
    def multi_sheet_xlsx(self, tmp_path):
        """Excel file with multiple sheets (entitlement + state)."""
        from openpyxl import Workbook

        wb = Workbook()

        # Entitlement sheet
        ws1 = wb.active
        ws1.title = "Entitlement"
        ws1.append(["NAME", "STA", "CDBG", "HOME"])
        ws1.append(["San Rafael", "CA", 523000, 187000])

        # State sheet
        ws2 = wb.create_sheet("State Programs")
        ws2.append(["NAME", "STA", "CDBG", "HOME"])
        ws2.append(["California", "CA", 50000000, 20000000])

        path = tmp_path / "test_multi.xlsx"
        wb.save(str(path))
        return str(path)

    def test_parse_extracts_correct_count(self, client, simple_xlsx):
        allocations = client._parse_excel(simple_xlsx, 2025)
        # 3 grantees × varying programs: San Rafael has 3 (CDBG+HOME+ESG),
        # Berkeley has 3, Portland has 2 (HOME is None) = 8 total
        assert len(allocations) == 8

    def test_parse_extracts_correct_amounts_in_cents(self, client, simple_xlsx):
        allocations = client._parse_excel(simple_xlsx, 2025)
        san_rafael_cdbg = [a for a in allocations
                           if a["grantee_name"] == "City of San Rafael"
                           and a["program"] == "CDBG"]
        assert len(san_rafael_cdbg) == 1
        assert san_rafael_cdbg[0]["amount_cents"] == 52300000  # $523,000 in cents

    def test_parse_skips_none_amounts(self, client, simple_xlsx):
        allocations = client._parse_excel(simple_xlsx, 2025)
        portland_home = [a for a in allocations
                         if a["grantee_name"] == "Portland"
                         and a["program"] == "HOME"]
        assert len(portland_home) == 0

    def test_parse_sets_fiscal_year(self, client, simple_xlsx):
        allocations = client._parse_excel(simple_xlsx, 2025)
        assert all(a["fiscal_year"] == 2025 for a in allocations)

    def test_parse_sets_state(self, client, simple_xlsx):
        allocations = client._parse_excel(simple_xlsx, 2025)
        ca_allocs = [a for a in allocations if a["state"] == "CA"]
        or_allocs = [a for a in allocations if a["state"] == "OR"]
        assert len(ca_allocs) == 6  # San Rafael (3) + Berkeley (3)
        assert len(or_allocs) == 2  # Portland (CDBG + ESG)

    def test_parse_normalizes_grantee_type_from_type_column(self, client, simple_xlsx):
        allocations = client._parse_excel(simple_xlsx, 2025)
        san_rafael = [a for a in allocations if a["grantee_name"] == "City of San Rafael"]
        assert all(a["grantee_type"] == "entitlement" for a in san_rafael)

    def test_parse_multi_sheet(self, client, multi_sheet_xlsx):
        allocations = client._parse_excel(multi_sheet_xlsx, 2025)
        entitlement = [a for a in allocations if a["grantee_type"] == "entitlement"]
        state = [a for a in allocations if a["grantee_type"] == "state"]
        assert len(entitlement) == 2  # San Rafael CDBG + HOME
        assert len(state) == 2  # California CDBG + HOME

    def test_parse_skips_total_rows(self, client, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Entitlement"
        ws.append(["NAME", "STA", "CDBG"])
        ws.append(["San Rafael", "CA", 100000])
        ws.append(["Total", "", 100000])
        ws.append(["Totals", "", 100000])

        path = tmp_path / "test_totals.xlsx"
        wb.save(str(path))

        allocations = client._parse_excel(str(path), 2025)
        grantee_names = [a["grantee_name"] for a in allocations]
        assert "Total" not in grantee_names
        assert "Totals" not in grantee_names
        assert len(allocations) == 1

    def test_parse_skips_zero_and_negative_amounts(self, client, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Entitlement"
        ws.append(["NAME", "STA", "CDBG"])
        ws.append(["City A", "CA", 0])
        ws.append(["City B", "CA", -5000])
        ws.append(["City C", "CA", 100000])

        path = tmp_path / "test_zeros.xlsx"
        wb.save(str(path))

        allocations = client._parse_excel(str(path), 2025)
        assert len(allocations) == 1
        assert allocations[0]["grantee_name"] == "City C"

    def test_parse_handles_dollar_formatted_amounts(self, client, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Entitlement"
        ws.append(["NAME", "STA", "CDBG"])
        ws.append(["City A", "CA", "$1,234,567.89"])

        path = tmp_path / "test_dollars.xlsx"
        wb.save(str(path))

        allocations = client._parse_excel(str(path), 2025)
        assert len(allocations) == 1
        # float("1234567.89") * 100 = 123456788.99... → int truncates to 123456788
        assert allocations[0]["amount_cents"] == 123456788

    def test_parse_skips_sheet_with_no_headers(self, client, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Notes"
        ws.append(["This is just a notes sheet"])
        ws.append(["No headers here"])

        ws2 = wb.create_sheet("Entitlement")
        ws2.append(["NAME", "STA", "CDBG"])
        ws2.append(["City A", "CA", 500000])

        path = tmp_path / "test_no_headers.xlsx"
        wb.save(str(path))

        allocations = client._parse_excel(str(path), 2025)
        assert len(allocations) == 1
        assert allocations[0]["grantee_name"] == "City A"

    def test_parse_sets_source_url_from_known_urls(self, client, simple_xlsx):
        allocations = client._parse_excel(simple_xlsx, 2025)
        expected_url = HUD_ALLOCATION_URLS[2025]
        assert all(a["source_url"] == expected_url for a in allocations)

    def test_parse_empty_source_url_for_unknown_year(self, client, simple_xlsx):
        allocations = client._parse_excel(simple_xlsx, 1999)
        assert all(a["source_url"] == "" for a in allocations)


# ==================== _load_allocations ====================


class TestLoadAllocations:
    """Tests for allocation loading with caching."""

    def test_returns_cached_data(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data
        result = client._load_allocations(2025)
        assert result is sample_allocations_data

    def test_returns_empty_for_unknown_year(self, client):
        result = client._load_allocations(1999)
        assert result == []

    def test_returns_empty_on_download_failure(self, client):
        with patch.object(client, "_download_excel", return_value=None):
            result = client._load_allocations(2025)
        assert result == []

    def test_populates_cache_on_success(self, client, tmp_path):
        fake_allocs = [{"grantee_name": "Test", "program": "CDBG"}]

        with patch.object(client, "_download_excel", return_value="/fake/path"), \
             patch.object(client, "_parse_excel", return_value=fake_allocs):
            result = client._load_allocations(2025)

        assert result == fake_allocs
        assert client._allocation_cache[2025] == fake_allocs


# ==================== get_allocations ====================


class TestGetAllocations:
    """Tests for the main get_allocations method."""

    def test_finds_grantee_by_exact_name(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.get_allocations("City of San Rafael", fiscal_year=2025)

        assert len(results) == 2
        programs = {r.program for r in results}
        assert programs == {"CDBG", "HOME"}

    def test_finds_grantee_by_partial_name(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        # Search "San Rafael" should match "City of San Rafael"
        results = client.get_allocations("San Rafael", fiscal_year=2025)

        assert len(results) == 2
        assert all(r.grantee_name == "City of San Rafael" for r in results)

    def test_filters_by_program(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.get_allocations("San Rafael", program="CDBG", fiscal_year=2025)

        assert len(results) == 1
        assert results[0].program == "CDBG"
        assert results[0].amount_cents == 52300000

    def test_program_filter_case_insensitive(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.get_allocations("San Rafael", program="cdbg", fiscal_year=2025)

        assert len(results) == 1
        assert results[0].program == "CDBG"

    def test_no_match_returns_empty(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.get_allocations("Nonexistent City", fiscal_year=2025)

        assert results == []

    def test_defaults_to_latest_year(self, client, sample_allocations_data):
        latest_year = max(HUD_ALLOCATION_URLS.keys())
        client._allocation_cache[latest_year] = sample_allocations_data

        results = client.get_allocations("San Rafael")

        assert len(results) == 2

    def test_returns_hudallocation_objects(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.get_allocations("San Rafael", fiscal_year=2025)

        for r in results:
            assert isinstance(r, HUDAllocation)
            assert r.state == "CA"
            assert r.fiscal_year == 2025
            assert r.grantee_type == "entitlement"

    def test_empty_allocations_returns_empty(self, client):
        client._allocation_cache[2025] = []

        results = client.get_allocations("San Rafael", fiscal_year=2025)

        assert results == []


# ==================== search_allocations ====================


class TestSearchAllocations:
    """Tests for state/program-based allocation search."""

    def test_filters_by_state_and_program(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.search_allocations(state="CA", program="CDBG", fiscal_year=2025)

        assert len(results) == 2
        names = [r.grantee_name for r in results]
        assert "City of San Rafael" in names
        assert "City of Berkeley" in names

    def test_sorted_by_amount_descending(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.search_allocations(state="CA", program="CDBG", fiscal_year=2025)

        assert results[0].amount_cents >= results[1].amount_cents
        assert results[0].grantee_name == "City of Berkeley"  # $3.2M > $523K

    def test_min_amount_filter(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.search_allocations(
            state="CA", program="CDBG", fiscal_year=2025,
            min_amount_cents=100000000,  # > $1M
        )

        assert len(results) == 1
        assert results[0].grantee_name == "City of Berkeley"

    def test_state_filter_case_insensitive(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.search_allocations(state="ca", program="CDBG", fiscal_year=2025)

        assert len(results) == 2

    def test_no_match_returns_empty(self, client, sample_allocations_data):
        client._allocation_cache[2025] = sample_allocations_data

        results = client.search_allocations(state="TX", program="CDBG", fiscal_year=2025)

        assert results == []

    def test_defaults_to_latest_year(self, client, sample_allocations_data):
        latest_year = max(HUD_ALLOCATION_URLS.keys())
        client._allocation_cache[latest_year] = sample_allocations_data

        results = client.search_allocations(state="OR", program="CDBG")

        assert len(results) == 1
        assert results[0].grantee_name == "Portland"


# ==================== get_available_* ====================


class TestAvailableMetadata:
    """Tests for available fiscal years and programs."""

    def test_fiscal_years_sorted_descending(self, client):
        years = client.get_available_fiscal_years()
        assert years == sorted(years, reverse=True)
        assert 2025 in years
        assert 2020 in years

    def test_programs_returns_all_six(self, client):
        programs = client.get_available_programs()
        assert len(programs) == 6
        assert programs["CDBG"] == "Community Development Block Grant"
        assert programs["HOME"] == "HOME Investment Partnerships Program"
        assert programs["ESG"] == "Emergency Solutions Grants"

    def test_programs_returns_copy(self, client):
        """Modifying returned dict should not affect original."""
        programs = client.get_available_programs()
        programs["NEW"] = "Fake Program"
        assert "NEW" not in HUD_PROGRAMS


# ==================== health ====================


class TestHealth:
    """Tests for health check."""

    def test_health_available_on_200(self, client):
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.headers = {"Content-Length": "12345"}

        with patch.object(client.session, "head", return_value=mock_response):
            status = client.health()

        assert status.is_available is True
        assert status.source_id == "hud_exchange-federal-US"
        assert status.source_type == "hud_exchange"
        assert status.errors == []
        assert status.metadata["latest_fiscal_year"] == max(HUD_ALLOCATION_URLS.keys())
        assert status.last_successful is not None

    def test_health_unavailable_on_error_status(self, client):
        mock_response = MagicMock()
        mock_response.status_code = 503

        with patch.object(client.session, "head", return_value=mock_response):
            status = client.health()

        assert status.is_available is False
        assert len(status.errors) == 1
        assert "503" in status.errors[0]
        assert status.last_successful is None

    def test_health_unavailable_on_exception(self, client):
        with patch.object(
            client.session, "head",
            side_effect=requests.ConnectionError("timeout"),
        ):
            status = client.health()

        assert status.is_available is False
        assert len(status.errors) == 1
        assert "timeout" in status.errors[0]

    def test_health_includes_cached_count(self, client, sample_allocations_data):
        latest_year = max(HUD_ALLOCATION_URLS.keys())
        client._allocation_cache[latest_year] = sample_allocations_data

        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.headers = {}

        with patch.object(client.session, "head", return_value=mock_response):
            status = client.health()

        assert status.available_count == 4

    def test_health_records_duration(self, client):
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.headers = {}

        with patch.object(client.session, "head", return_value=mock_response):
            status = client.health()

        assert status.check_duration_ms >= 0


# ==================== validate ====================


class TestValidate:
    """Tests for preflight validation."""

    def test_validate_success(self, client):
        mock_response = MagicMock()
        mock_response.status_code = 200

        with patch.object(client.session, "head", return_value=mock_response):
            result = client.validate()

        assert result.is_valid is True
        assert result.config_valid is True
        assert result.api_reachable is True
        assert result.errors == []

    def test_validate_403_still_reachable(self, client):
        mock_response = MagicMock()
        mock_response.status_code = 403

        with patch.object(client.session, "head", return_value=mock_response):
            result = client.validate()

        assert result.is_valid is True
        assert result.api_reachable is True
        assert len(result.warnings) == 1
        assert "block" in result.warnings[0].lower()

    def test_validate_server_error(self, client):
        mock_response = MagicMock()
        mock_response.status_code = 500

        with patch.object(client.session, "head", return_value=mock_response):
            result = client.validate()

        assert result.is_valid is False
        assert result.api_reachable is False
        assert len(result.errors) == 1
        assert "500" in result.errors[0]

    def test_validate_network_error(self, client):
        with patch.object(
            client.session, "head",
            side_effect=requests.ConnectionError("DNS failed"),
        ):
            result = client.validate()

        assert result.is_valid is False
        assert result.api_reachable is False
        assert "DNS failed" in result.errors[0]

    def test_validate_missing_openpyxl(self, client):
        with patch.dict("sys.modules", {"openpyxl": None}):
            # Force ImportError on openpyxl import
            with patch("builtins.__import__", side_effect=_openpyxl_import_blocker):
                result = client.validate()

        assert result.config_valid is False
        assert any("openpyxl" in e for e in result.errors)

    def test_validate_records_duration(self, client):
        mock_response = MagicMock()
        mock_response.status_code = 200

        with patch.object(client.session, "head", return_value=mock_response):
            result = client.validate()

        assert result.check_duration_ms >= 0


# ==================== Storage Mappers ====================


class TestHudAllocationToStorage:
    """Tests for mapping HUDAllocation to storage format."""

    def test_maps_all_fields(self, sample_allocation):
        result = hud_allocation_to_storage(sample_allocation, "city-san-rafael")

        assert result["program_id"] == "cdbg"
        assert result["jurisdiction_id"] == "city-san-rafael"
        assert result["fiscal_year"] == "2025"
        assert result["allocation_amount_cents"] == 52300000
        assert result["allocation_status"] == "allocated"
        assert result["administering_entity"] == "City of San Rafael"
        assert result["grantee_type"] == "entitlement"
        assert "FY2025" in result["source_url"]
        assert result["metadata"]["state"] == "CA"
        assert result["metadata"]["source"] == "hud_cpd_formula_allocations"

    def test_program_id_is_lowercase(self, sample_allocation):
        result = hud_allocation_to_storage(sample_allocation, "city-san-rafael")
        assert result["program_id"] == "cdbg"

    def test_fiscal_year_is_string(self, sample_allocation):
        result = hud_allocation_to_storage(sample_allocation, "city-san-rafael")
        assert isinstance(result["fiscal_year"], str)
        assert result["fiscal_year"] == "2025"


class TestExtractAllocationsToStorage:
    """Tests for multi-year extraction to storage."""

    def test_extracts_across_years(self):
        mock_client = MagicMock(spec=HUDExchangeClient)
        alloc = HUDAllocation(
            grantee_name="San Rafael",
            state="CA",
            program="CDBG",
            fiscal_year=2025,
            amount_cents=52300000,
            grantee_type="entitlement",
            source_url="https://example.com",
        )
        mock_client.get_allocations.return_value = [alloc]
        mock_client.get_available_fiscal_years.return_value = [2025, 2024]

        results = extract_allocations_to_storage(
            mock_client, "San Rafael", "city-san-rafael"
        )

        # 2 years × 1 allocation each = 2
        assert len(results) == 2
        assert results[0]["program_id"] == "cdbg"
        assert results[0]["jurisdiction_id"] == "city-san-rafael"
        assert results[0]["allocation_amount_cents"] == 52300000

    def test_specific_fiscal_years(self):
        mock_client = MagicMock(spec=HUDExchangeClient)
        mock_client.get_allocations.return_value = []

        results = extract_allocations_to_storage(
            mock_client, "San Rafael", "city-san-rafael",
            fiscal_years=[2023],
        )

        assert results == []
        mock_client.get_allocations.assert_called_once_with("San Rafael", fiscal_year=2023)

    def test_empty_when_no_allocations(self):
        mock_client = MagicMock(spec=HUDExchangeClient)
        mock_client.get_allocations.return_value = []
        mock_client.get_available_fiscal_years.return_value = [2025]

        results = extract_allocations_to_storage(
            mock_client, "Nowhere", "city-nowhere"
        )

        assert results == []


# ==================== create_hud_exchange_client ====================


class TestCreateClient:
    """Tests for the factory function."""

    def test_creates_client_with_defaults(self):
        c = create_hud_exchange_client()
        assert c.jurisdiction_id == "federal-US"
        assert c.request_delay == 1.0
        assert c.platform_name == "hud_exchange"


# ==================== Module Constants ====================


class TestModuleConstants:
    """Tests for module-level constants."""

    def test_allocation_urls_are_all_xlsx(self):
        for year, url in HUD_ALLOCATION_URLS.items():
            assert url.endswith(".xlsx"), f"FY{year} URL doesn't end with .xlsx"
            assert f"FY{year}" in url, f"FY{year} URL doesn't contain year"

    def test_allocation_urls_cover_recent_years(self):
        assert 2025 in HUD_ALLOCATION_URLS
        assert 2020 in HUD_ALLOCATION_URLS

    def test_hud_programs_has_six_entries(self):
        assert len(HUD_PROGRAMS) == 6
        for code in ["CDBG", "HOME", "ESG", "HOPWA", "HTF", "RHP"]:
            assert code in HUD_PROGRAMS

    def test_column_mappings_cover_all_fields(self):
        expected_fields = {"grantee", "state", "type", "cdbg", "home", "esg", "hopwa", "htf", "rhp"}
        assert set(COLUMN_MAPPINGS.keys()) == expected_fields


# ==================== Helper ====================


def _openpyxl_import_blocker(name, *args, **kwargs):
    """Import blocker that only blocks openpyxl."""
    if name == "openpyxl" or (hasattr(name, "startswith") and name.startswith("openpyxl")):
        raise ImportError("openpyxl not installed")
    return original_import(name, *args, **kwargs)


import builtins
original_import = builtins.__import__
