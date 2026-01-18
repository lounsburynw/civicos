"""
HUD Exchange / HUD CPD Allocation Data Client

Retrieves federal housing program allocations (CDBG, HOME, ESG, HOPWA, HTF, RHP)
from HUD's official data sources.

Data Sources:
- HUD CPD Formula Allocations Excel files (primary)
- HUD Exchange Awards & Allocations search (fallback)

Usage:
    client = HUDExchangeClient()
    allocations = client.get_allocations("San Rafael", "CDBG", 2025)
    # Returns: [{"jurisdiction": "San Rafael", "program": "CDBG", "amount_cents": 123456700, ...}]

The client downloads official HUD Excel spreadsheets which contain allocation
data for all entitlement grantees. This approach is more reliable than scraping
the HUD Exchange website, which requires JavaScript rendering.
"""

import logging
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional

import requests

from civicos_extraction.clients.base import HealthStatus, ValidationResult

logger = logging.getLogger(__name__)


# Known HUD CPD allocation file URLs by fiscal year
# URL pattern: https://www.hud.gov/sites/dfiles/CPD/documents/FY{year}-Formula-Allocations-All-Grantees.xlsx
HUD_ALLOCATION_URLS = {
    2025: "https://www.hud.gov/sites/dfiles/CPD/documents/FY2025-Formula-Allocations-All-Grantees.xlsx",
    2024: "https://www.hud.gov/sites/dfiles/CPD/documents/FY2024-Formula-Allocations-All-Grantees.xlsx",
    2023: "https://www.hud.gov/sites/dfiles/CPD/documents/FY2023-Formula-Allocations-All-Grantees.xlsx",
    2022: "https://www.hud.gov/sites/dfiles/CPD/documents/FY2022-Formula-Allocations-All-Grantees.xlsx",
    2021: "https://www.hud.gov/sites/dfiles/CPD/documents/FY2021-Formula-Allocations-All-Grantees.xlsx",
    2020: "https://www.hud.gov/sites/dfiles/CPD/documents/FY2020-Formula-Allocations-All-Grantees.xlsx",
}

# HUD program codes and their full names
HUD_PROGRAMS = {
    "CDBG": "Community Development Block Grant",
    "HOME": "HOME Investment Partnerships Program",
    "ESG": "Emergency Solutions Grants",
    "HOPWA": "Housing Opportunities for Persons With AIDS",
    "HTF": "Housing Trust Fund",
    "RHP": "CDBG Recovery Housing Program",
}

# Column name mappings for different fiscal years (column names may vary)
# The FY2025 file uses: NAME, STA, Type, CDBG, RHP, HOME, ESG, HOPWA, HTF
COLUMN_MAPPINGS = {
    "grantee": ["NAME", "Grantee Name", "Grantee", "Name", "Recipient Name"],
    "state": ["STA", "STATE", "State", "ST", "State Code"],
    "type": ["Type", "Grantee Type"],
    "cdbg": ["CDBG", "CDBG Allocation", "CDBG Amount"],
    "home": ["HOME", "HOME Allocation", "HOME Amount"],
    "esg": ["ESG", "ESG Allocation", "ESG Amount"],
    "hopwa": ["HOPWA", "HOPWA Allocation", "HOPWA Amount"],
    "htf": ["HTF", "HTF Allocation", "HTF Amount"],
    "rhp": ["RHP", "RHP Allocation", "RHP Amount"],
}


@dataclass
class HUDAllocation:
    """
    A single HUD allocation record.

    Represents an allocation of federal funds to a specific grantee
    for a specific program and fiscal year.
    """

    grantee_name: str  # "City of San Rafael", "San Rafael"
    state: str  # "CA"
    program: str  # "CDBG", "HOME", etc.
    fiscal_year: int  # 2025
    amount_cents: int  # Amount in cents (dollars * 100)
    grantee_type: str  # "entitlement", "state", "insular_area"
    source_url: str  # URL of the data source

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for storage."""
        return {
            "grantee_name": self.grantee_name,
            "state": self.state,
            "program": self.program,
            "fiscal_year": self.fiscal_year,
            "amount_cents": self.amount_cents,
            "grantee_type": self.grantee_type,
            "source_url": self.source_url,
        }


class HUDExchangeClient:
    """
    HUD Exchange / HUD CPD allocation data client.

    Retrieves allocation data from HUD's official Excel spreadsheets
    which are published annually with formula program allocations.

    Features:
    - Downloads official HUD Excel files (no scraping required)
    - Caches downloaded files to avoid re-downloading
    - Normalizes grantee names for matching
    - Supports CDBG, HOME, ESG, HOPWA, HTF, RHP programs
    """

    def __init__(
        self,
        jurisdiction_id: str = "federal-US",
        cache_dir: Optional[str] = None,
        request_delay: float = 1.0,
    ):
        """
        Initialize HUD Exchange client.

        Args:
            jurisdiction_id: Civic jurisdiction ID (default "federal-US")
            cache_dir: Directory to cache downloaded Excel files (default: temp dir)
            request_delay: Delay between requests in seconds (default 1.0)
        """
        self.jurisdiction_id = jurisdiction_id
        self.cache_dir = cache_dir or tempfile.gettempdir()
        self.request_delay = request_delay
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        })
        self.last_request_time = 0.0
        self._allocation_cache: Dict[int, List[Dict[str, Any]]] = {}

    @property
    def platform_name(self) -> str:
        return "hud_exchange"

    @property
    def source_id(self) -> str:
        """Unique identifier: platform-jurisdiction."""
        return f"hud_exchange-{self.jurisdiction_id}"

    @property
    def source_type(self) -> str:
        """Type of source."""
        return "hud_exchange"

    def _throttle_request(self) -> None:
        """Ensure minimum delay between requests."""
        elapsed = time.time() - self.last_request_time
        if elapsed < self.request_delay:
            time.sleep(self.request_delay - elapsed)
        self.last_request_time = time.time()

    def _download_excel(self, url: str, fiscal_year: int) -> Optional[str]:
        """
        Download Excel file from HUD and save to cache.

        Args:
            url: URL of the Excel file
            fiscal_year: Fiscal year for cache key

        Returns:
            Path to downloaded file, or None if download failed
        """
        import os

        cache_path = os.path.join(
            self.cache_dir,
            f"hud_cpd_fy{fiscal_year}.xlsx"
        )

        # Use cached file if it exists and is recent (< 7 days old)
        if os.path.exists(cache_path):
            file_age = time.time() - os.path.getmtime(cache_path)
            if file_age < 7 * 24 * 60 * 60:  # 7 days
                logger.debug(f"Using cached HUD file: {cache_path}")
                return cache_path

        self._throttle_request()

        try:
            response = self.session.get(url, timeout=60)
            response.raise_for_status()

            with open(cache_path, "wb") as f:
                f.write(response.content)

            logger.info(f"Downloaded HUD allocation file: {cache_path}")
            return cache_path

        except requests.RequestException as e:
            logger.error(f"Failed to download HUD Excel file: {e}")
            return None

    def _parse_excel(
        self,
        file_path: str,
        fiscal_year: int,
    ) -> List[Dict[str, Any]]:
        """
        Parse HUD Excel file and extract allocation data.

        Args:
            file_path: Path to downloaded Excel file
            fiscal_year: Fiscal year of the data

        Returns:
            List of allocation records
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required for HUDExchangeClient. "
                "Install with: pip install openpyxl"
            )

        allocations: List[Dict[str, Any]] = []
        source_url = HUD_ALLOCATION_URLS.get(fiscal_year, "")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            # Process each sheet (typically one per grantee type)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Determine grantee type from sheet name
                grantee_type = self._infer_grantee_type(sheet_name)

                # Find header row and column indices
                headers: Dict[str, int] = {}
                header_row = None

                for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
                    # Look for header row - must have NAME or Grantee, plus CDBG or HOME
                    if row:
                        row_strs = [str(cell).upper() if cell else "" for cell in row]
                        has_name = any("NAME" in s or "GRANTEE" in s for s in row_strs)
                        has_program = any(p in row_strs for p in ["CDBG", "HOME", "ESG"])
                        if has_name and has_program:
                            for col_idx, cell in enumerate(row):
                                if cell:
                                    headers[str(cell).strip()] = col_idx
                            header_row = row_idx
                            break

                if not headers:
                    logger.debug(f"No headers found in sheet: {sheet_name}")
                    continue

                # Map columns to our expected fields
                col_map = self._map_columns(headers)

                if "grantee" not in col_map:
                    logger.debug(f"No grantee column found in sheet: {sheet_name}")
                    continue

                # Parse data rows
                for row in sheet.iter_rows(
                    min_row=header_row + 1 if header_row else 2,
                    values_only=True
                ):
                    if not row or not row[col_map["grantee"]]:
                        continue

                    grantee_name = str(row[col_map["grantee"]]).strip()
                    if not grantee_name or grantee_name.lower() in ["total", "totals", ""]:
                        continue

                    state = ""
                    if "state" in col_map and row[col_map["state"]]:
                        state = str(row[col_map["state"]]).strip()

                    # Get grantee type from row if Type column exists, else use sheet-inferred type
                    row_grantee_type = grantee_type
                    if "type" in col_map and row[col_map["type"]]:
                        row_grantee_type = self._normalize_grantee_type(str(row[col_map["type"]]).strip())

                    # Extract allocations for each program
                    for program in ["cdbg", "home", "esg", "hopwa", "htf", "rhp"]:
                        if program not in col_map:
                            continue

                        amount = row[col_map[program]]
                        if amount is None or amount == "":
                            continue

                        # Convert to cents (amounts are in dollars)
                        try:
                            # Convert any cell value to string first for safe parsing
                            amount_str = str(amount).replace(",", "").replace("$", "").strip()
                            amount_cents = int(float(amount_str) * 100)
                        except (ValueError, TypeError):
                            continue

                        if amount_cents <= 0:
                            continue

                        allocations.append({
                            "grantee_name": grantee_name,
                            "state": state,
                            "program": program.upper(),
                            "fiscal_year": fiscal_year,
                            "amount_cents": amount_cents,
                            "grantee_type": row_grantee_type,
                            "source_url": source_url,
                        })

            wb.close()
            logger.info(
                f"Parsed {len(allocations)} allocations from FY{fiscal_year} file"
            )

        except Exception as e:
            logger.error(f"Error parsing HUD Excel file: {e}")
            raise

        return allocations

    def _infer_grantee_type(self, sheet_name: str) -> str:
        """Infer grantee type from sheet name."""
        sheet_lower = sheet_name.lower()
        if "entitlement" in sheet_lower:
            return "entitlement"
        elif "state" in sheet_lower:
            return "state"
        elif "insular" in sheet_lower:
            return "insular_area"
        elif "urban county" in sheet_lower or "consortium" in sheet_lower:
            return "entitlement"
        else:
            return "entitlement"

    def _normalize_grantee_type(self, type_str: str) -> str:
        """Normalize grantee type from Excel Type column."""
        type_lower = type_str.lower()
        if "non-entitlement" in type_lower or "state" in type_lower:
            return "state"
        elif "principal city" in type_lower:
            return "entitlement"
        elif "metro city" in type_lower:
            return "entitlement"
        elif "urban county" in type_lower:
            return "entitlement"
        elif "consortium" in type_lower:
            return "entitlement"
        elif "insular" in type_lower:
            return "insular_area"
        else:
            return "entitlement"

    def _map_columns(self, headers: Dict[str, int]) -> Dict[str, int]:
        """Map Excel column names to expected field names."""
        col_map: Dict[str, int] = {}

        for field, possible_names in COLUMN_MAPPINGS.items():
            for header_name, col_idx in headers.items():
                header_upper = header_name.upper().strip()
                for possible in possible_names:
                    # Exact match (case-insensitive)
                    if possible.upper() == header_upper:
                        col_map[field] = col_idx
                        break
                if field in col_map:
                    break

        return col_map

    def _normalize_grantee_name(self, name: str) -> str:
        """Normalize grantee name for matching."""
        name = name.lower().strip()
        # Remove common prefixes
        for prefix in ["city of ", "town of ", "county of ", "village of "]:
            if name.startswith(prefix):
                name = name[len(prefix):]
        return name

    def _load_allocations(self, fiscal_year: int) -> List[Dict[str, Any]]:
        """Load allocations for a fiscal year, using cache if available."""
        if fiscal_year in self._allocation_cache:
            return self._allocation_cache[fiscal_year]

        if fiscal_year not in HUD_ALLOCATION_URLS:
            logger.warning(f"No HUD allocation URL for FY{fiscal_year}")
            return []

        file_path = self._download_excel(
            HUD_ALLOCATION_URLS[fiscal_year],
            fiscal_year
        )

        if not file_path:
            return []

        allocations = self._parse_excel(file_path, fiscal_year)
        self._allocation_cache[fiscal_year] = allocations
        return allocations

    def health(self) -> HealthStatus:
        """
        Check source availability and return standardized status.

        Performs a lightweight check by attempting to access the latest
        allocation file.
        """
        start_time = time.time()
        errors: List[str] = []
        is_available = False
        available_count = 0
        metadata: Dict[str, Any] = {}

        try:
            # Check if we can reach the latest allocation file
            latest_year = max(HUD_ALLOCATION_URLS.keys())
            url = HUD_ALLOCATION_URLS[latest_year]

            self._throttle_request()
            response = self.session.head(url, timeout=30)

            if response.status_code == 200:
                is_available = True
                metadata["latest_fiscal_year"] = latest_year
                metadata["content_length"] = response.headers.get("Content-Length")

                # Quick count: load cached data if available
                if latest_year in self._allocation_cache:
                    available_count = len(self._allocation_cache[latest_year])
            else:
                errors.append(f"HUD server returned status {response.status_code}")

        except Exception as e:
            errors.append(f"Health check error: {str(e)}")
            logger.warning(
                "Health check failed",
                extra={
                    "error": str(e),
                    "platform": self.platform_name,
                },
            )

        check_duration_ms = (time.time() - start_time) * 1000
        now = datetime.now()

        return HealthStatus(
            source_id=self.source_id,
            source_type=self.source_type,
            jurisdiction_id=self.jurisdiction_id,
            is_available=is_available,
            available_count=available_count,
            last_checked=now,
            check_duration_ms=round(check_duration_ms, 2),
            errors=errors,
            last_successful=now if is_available else None,
            metadata=metadata,
        )

    def validate(self) -> ValidationResult:
        """Validate source configuration and accessibility."""
        start_time = time.time()
        errors: List[str] = []
        warnings: List[str] = []
        config_valid = True
        api_reachable = False
        metadata: Dict[str, Any] = {}

        # Check openpyxl is available
        try:
            from openpyxl import load_workbook as _load  # noqa: F401
            del _load
        except ImportError:
            errors.append(
                "openpyxl not installed. Run: pip install openpyxl"
            )
            config_valid = False

        # Check if we can reach HUD server
        if config_valid:
            try:
                latest_year = max(HUD_ALLOCATION_URLS.keys())
                url = HUD_ALLOCATION_URLS[latest_year]

                self._throttle_request()
                response = self.session.head(url, timeout=30)

                if response.status_code == 200:
                    api_reachable = True
                    metadata["latest_fiscal_year"] = latest_year
                elif response.status_code == 403:
                    warnings.append(
                        "HUD server may block automated requests. "
                        "Some fiscal years may require manual download."
                    )
                    api_reachable = True  # HEAD blocked but GET might work
                else:
                    errors.append(f"Cannot reach HUD server: HTTP {response.status_code}")

            except requests.RequestException as e:
                errors.append(f"Network error: {str(e)}")

        check_duration_ms = (time.time() - start_time) * 1000

        return ValidationResult(
            is_valid=len(errors) == 0,
            config_valid=config_valid,
            api_reachable=api_reachable,
            errors=errors,
            warnings=warnings,
            check_duration_ms=round(check_duration_ms, 2),
            metadata=metadata,
        )

    def get_allocations(
        self,
        grantee_name: str,
        program: Optional[str] = None,
        fiscal_year: Optional[int] = None,
    ) -> List[HUDAllocation]:
        """
        Get allocations for a specific grantee.

        Args:
            grantee_name: Name of the grantee (e.g., "San Rafael", "City of San Rafael")
            program: Optional program filter ("CDBG", "HOME", etc.)
            fiscal_year: Optional fiscal year (default: latest available)

        Returns:
            List of HUDAllocation objects
        """
        if fiscal_year is None:
            fiscal_year = max(HUD_ALLOCATION_URLS.keys())

        allocations = self._load_allocations(fiscal_year)
        if not allocations:
            return []

        # Normalize search name
        search_name = self._normalize_grantee_name(grantee_name)

        results: List[HUDAllocation] = []
        for alloc in allocations:
            # Match by normalized name (partial match)
            alloc_name = self._normalize_grantee_name(alloc["grantee_name"])
            if search_name not in alloc_name and alloc_name not in search_name:
                continue

            # Filter by program if specified
            if program and alloc["program"] != program.upper():
                continue

            results.append(HUDAllocation(
                grantee_name=alloc["grantee_name"],
                state=alloc["state"],
                program=alloc["program"],
                fiscal_year=alloc["fiscal_year"],
                amount_cents=alloc["amount_cents"],
                grantee_type=alloc["grantee_type"],
                source_url=alloc["source_url"],
            ))

        logger.info(
            f"Found {len(results)} allocations for '{grantee_name}' in FY{fiscal_year}"
        )
        return results

    def search_allocations(
        self,
        state: str = "CA",
        program: str = "CDBG",
        fiscal_year: Optional[int] = None,
        min_amount_cents: int = 0,
    ) -> List[HUDAllocation]:
        """
        Search allocations by state and program.

        Args:
            state: State code (e.g., "CA")
            program: Program name (e.g., "CDBG", "HOME")
            fiscal_year: Fiscal year (default: latest available)
            min_amount_cents: Minimum allocation amount to include

        Returns:
            List of HUDAllocation objects sorted by amount (descending)
        """
        if fiscal_year is None:
            fiscal_year = max(HUD_ALLOCATION_URLS.keys())

        allocations = self._load_allocations(fiscal_year)
        if not allocations:
            return []

        results: List[HUDAllocation] = []
        for alloc in allocations:
            if alloc["state"].upper() != state.upper():
                continue

            if alloc["program"] != program.upper():
                continue

            if alloc["amount_cents"] < min_amount_cents:
                continue

            results.append(HUDAllocation(
                grantee_name=alloc["grantee_name"],
                state=alloc["state"],
                program=alloc["program"],
                fiscal_year=alloc["fiscal_year"],
                amount_cents=alloc["amount_cents"],
                grantee_type=alloc["grantee_type"],
                source_url=alloc["source_url"],
            ))

        # Sort by amount descending
        results.sort(key=lambda x: x.amount_cents, reverse=True)

        logger.info(
            f"Found {len(results)} {program} allocations in {state} for FY{fiscal_year}"
        )
        return results

    def get_available_fiscal_years(self) -> List[int]:
        """Get list of available fiscal years."""
        return sorted(HUD_ALLOCATION_URLS.keys(), reverse=True)

    def get_available_programs(self) -> Dict[str, str]:
        """Get dict of available programs (code -> full name)."""
        return HUD_PROGRAMS.copy()


def create_hud_exchange_client() -> HUDExchangeClient:
    """
    Create a HUD Exchange client configured for federal housing programs.

    Returns:
        Configured HUDExchangeClient
    """
    return HUDExchangeClient()


# ==================== Storage Mappers ====================


def hud_allocation_to_storage(
    allocation: HUDAllocation,
    jurisdiction_id: str,
) -> Dict[str, Any]:
    """
    Map HUD allocation to federal_program_allocations storage format.

    Args:
        allocation: HUDAllocation from HUDExchangeClient
        jurisdiction_id: Civic jurisdiction ID (e.g., "city-san-rafael")

    Returns:
        Dict ready for store_federal_program_allocations()
    """
    # Map HUD program to our program_id format
    program_id = allocation.program.lower()  # "cdbg", "home", etc.

    return {
        "program_id": program_id,
        "jurisdiction_id": jurisdiction_id,
        "fiscal_year": str(allocation.fiscal_year),
        "allocation_amount_cents": allocation.amount_cents,
        "allocation_status": "allocated",
        "administering_entity": allocation.grantee_name,
        "grantee_type": allocation.grantee_type,
        "source_url": allocation.source_url,
        "metadata": {
            "state": allocation.state,
            "source": "hud_cpd_formula_allocations",
        },
    }


def extract_allocations_to_storage(
    client: HUDExchangeClient,
    grantee_name: str,
    jurisdiction_id: str,
    fiscal_years: Optional[List[int]] = None,
) -> List[Dict[str, Any]]:
    """
    Extract allocations for a grantee across multiple fiscal years.

    Args:
        client: HUDExchangeClient instance
        grantee_name: Name of the grantee (e.g., "San Rafael")
        jurisdiction_id: Civic jurisdiction ID
        fiscal_years: List of fiscal years to extract (default: all available)

    Returns:
        List of allocation dicts ready for storage
    """
    if fiscal_years is None:
        fiscal_years = client.get_available_fiscal_years()

    results: List[Dict[str, Any]] = []

    for fy in fiscal_years:
        allocations = client.get_allocations(grantee_name, fiscal_year=fy)
        for alloc in allocations:
            results.append(hud_allocation_to_storage(alloc, jurisdiction_id))

    logger.info(
        f"Extracted {len(results)} allocations for {jurisdiction_id} "
        f"across {len(fiscal_years)} fiscal years"
    )
    return results
